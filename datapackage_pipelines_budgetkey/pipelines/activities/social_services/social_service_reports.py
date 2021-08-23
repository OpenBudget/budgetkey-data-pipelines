import dataflows as DF

# משרד	מינהל	אגף	שם השירות	תיאור השירות
# אוכלוסיית יעד	קבוצת גיל	תחום התערבות	אופן התערבות
# ניתן מכוח	קישור	מספר קטלוגי (אם רלוונטי)
services_query = '''
    select
        office as "משרד",
        unit as "מינהל",
        subunit as "אגף",
        name as "שם השירות",
        description as "תיאור השירות",
        target_audience as "אוכלוסיית יעד",
        target_age_group as "קבוצת גיל",
        subject as "תחום התערבות",
        intervention as "אופן התערבות",
        catalog_number as "מספר קטלוגי (אם רלוונטי)"
    from activities
    where office{office}
    order by 1, 2, 3, 4
'''

# שנה	משרד	מינהל	אגף	שם השירות	מספר מוטבים	תקנה תקציבית	תקציב מאושר	תקציב ביצוע בפועל
budget_query = '''
    with beneficiaries as (select id, jsonb_array_elements(beneficiaries) as beneficiaries from activities where office{office} and beneficiaries is not null and beneficiaries != 'null'::jsonb),
         manual_budgets as (select id, jsonb_array_elements("manualBudget") as manual_budgets from activities where office{office} and "manualBudget" is not null and "manualBudget" != 'null'::jsonb),
         budget_items as (select id, jsonb_array_elements("budgetItems") as budget_items from activities where office{office} and "budgetItems" is not null and "budgetItems" != 'null'::jsonb),
         years1 as (
             select
                coalesce(beneficiaries.id, manual_budgets.id) as id, 
                coalesce(beneficiaries->'year', manual_budgets->'year') as year,
                beneficiaries, manual_budgets
                from beneficiaries
                full join manual_budgets on (beneficiaries.id=manual_budgets.id AND beneficiaries->'year'=manual_budgets->'year')
        ),
        years as (
            select
                coalesce(years1.id, budget_items.id) as id, 
                coalesce(year, budget_items->'year') as year,
                beneficiaries, manual_budgets, budget_items
            from years1
            full join budget_items on (years1.id=budget_items.id AND year=budget_items->'year')
        )
    select
        year as "שנה",
        office as "משרד",
        unit as "מינהל",
        subunit as "אגף",
        name as "שם השירות",
        years.beneficiaries->'num_beneficiaries' as "מספר מוטבים",
        array_agg(years.budget_items->'code') as "תקנה תקציבית",
        years.manual_budgets->'approved' as "תקציב מאושר",
        years.manual_budgets->'executed' as "תקציב ביצוע בפועל"
    from activities
    join years on (activities.id=years.id)
    where office{office}
    group by 1,2,3,4,5,6,8,9
    order by 2, 3, 4, 5, 1 desc
'''

# שנה	משרד	מינהל	אגף	שם השירות	ח.פ מפעיל	שם מפעיל	אזור פעילות
suppliers_query = '''
    with suppliers as (
        select 
            id,
            jsonb_array_elements(suppliers) as supplier
        from activities
        where office{office} and suppliers is not null and suppliers != 'null'::jsonb),
    suppliers_years as (
        select
            id, supplier, jsonb_array_elements(supplier->'activity_years') as year
        from suppliers
    )
    select
        year as "שנה",
        office as "משרד",
        unit as "מינהל",
        subunit as "אגף",
        name as "שם השירות",
        supplier->'entity_id' as "ח.פ מפעיל",
        supplier->'entity_name' as "שם מפעיל",
        supplier->'geo' as "אזור פעילות"
    from activities
    join suppliers_years on (activities.id=suppliers_years.id)
    where office{office}
    order by 2, 3, 4, 5, 1 desc    
'''

# שנה	משרד	מינהל	אגף	שם השירות
# 	מספר מכרז	סוג הליך תחרותי	תוקף	מספר אופציות	סך משך האופציות	מודל תמחור
tenders_query = '''
    with tenders as (
        select 
            id,
            jsonb_array_elements(tenders) as tender
        from activities
        where office{office} and tenders is not null and tenders != 'null'::jsonb)
    select
        office as "משרד",
        unit as "מינהל",
        subunit as "אגף",
        name as "שם השירות",
        tender->'tender_id' as "מספר מכרז",
        tender->'tender_type_he' as "סוג הליך תחרותי",
        tender->'end_date' as "תוקף",
        tender->'option_num' as "מספר אופציות",
        tender->'option_duration' as "סך משך האופציות",
        case tender->>'pricing_model'
            when 'fixed' then 'מחיר קבוע'
            when 'proposal' then 'הצעת מחיר'
            when 'combined' then 'משולב'
        else
            null
        end "מודל תמחור"
    from activities
    join tenders on (activities.id=tenders.id)
    where office{office}
    order by 2, 3, 4, 5, 1 desc    

'''

base_path = '/var/datapackages/reports/services/'
report_kinds = [
    dict(name='services', filename='שירותים חברתיים - {office}.xlsx', sheet='שירות ומאפיינים', query=services_query),
    dict(name='budget', filename='שירותים חברתיים - {office}.xlsx', sheet='תקציב ומוטבים', query=budget_query),
    dict(name='suppliers', filename='שירותים חברתיים - {office}.xlsx', sheet='מפעילים', query=suppliers_query),
    dict(name='tenders', filename='שירותים חברתיים - {office}.xlsx', sheet='הליכי רכש', query=tenders_query),
]
offices = [
    dict(name='משרד הבריאות', selector="='משרד הבריאות'"),
    dict(name='משרד החינוך', selector="='משרד החינוך'"),
    dict(name='משרד הרווחה', selector="='משרד הרווחה'"),
]

def do_query():
    def func(row):
        office = row['office']
        kind = row['kind']
        filename = kind['filename'].format(office=office['name'])
        sheetname = kind['sheet']
        query = kind['query'].format(office=office['selector'])

        import openpyxl

        try:
            wb = openpyxl.load_workbook(base_path + filename)
        except:
            wb = openpyxl.Workbook()
            for s in wb.worksheets:
                wb.remove(s)
        try:
            ws = wb[sheetname]
        except:
            ws = wb.create_sheet(sheetname)

        ws.sheet_view.rightToleft = True

        for i, _ in enumerate(ws.columns, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 90
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].bestFit = True

        wb.save(base_path + filename)
        wb.close()
        del wb

        DF.Flow(
            DF.load('env://DPP_DB_ENGINE', query=query),
            DF.update_resource(-1, name=sheetname, path=filename),
            DF.printer(),
            DF.dump_to_path(base_path, format='excel',
                options=dict(
                    update_existing=base_path + filename,
                    sheetname=sheetname,
                )
            )
        ).process()

    return func


def flow(*_):
    return DF.Flow(
        [dict(office=office, kind=kind) for office in offices for kind in report_kinds],
        do_query(),
        DF.printer(),
        DF.update_resource(-1, **{'dpp:streaming': True}),
    )


if __name__ == '__main__':
    flow().process()